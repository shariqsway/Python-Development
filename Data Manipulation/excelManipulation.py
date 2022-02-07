# Import the package
import openpyxl

# When working with spreadsheets we need to start with an workbook object.
# We can either create an empty workbook in memory or load an existing workbook.

# Create a new workbook.
# wb = openpyxl.Workbook()

# Load existing workbook.
wb = openpyxl.load_workbook("transactions.xlsx")

# Returns a list with the names of each sheet in our workbook.
print(wb.sheetnames)

# To access a sheet in the workbook.
sheet_1 = wb["Sheet1"]

# Create a new sheet.
sheet_2 = wb.create_sheet(title="Sheet2", index=0)

# To reference a single cell
cell_a1 = sheet_1["a1"]

# Return the content in that cell. In this case it will return "transaction_id"
print(cell_a1.value)

# Return the row number. Returns "1" here
print(cell_a1.row)

# Return the column number. Returns "1", not "A"
print(cell_a1.column)

# Returns the column letter. Returns "A".
print(cell_a1.column_letter)

# Return the coordinate of the cell. Returns "A1".
print(cell_a1.coordinate)

# Returns single cell.
cell_b2 = sheet_1.cell(row=1, column=2)
print(cell_b2.value)

# Get number of columns.
print(sheet_1.max_column)

# Get number of rows.
print(sheet_1.max_row)

# Iterate rows and columns and get each value.
for row in range(1, sheet_1.max_row + 1):
    for column in range(1, sheet_1.max_column + 1):
        cell = sheet_1.cell(row, column)
        print(cell.value)

# To access a column in the sheet.
column_a = sheet_1["a"]
print(column_a)

# To access range of columns.
cell_a_c = sheet_1["a:c"]
print(cell_a_c)

# To access single row.
cell_2_4 = sheet_1["2"]
print(cell_2_4)

# To access range of rows.
cell_2_4 = sheet_1["2:4"]
print(cell_2_4)

# To access range of cells.
cell_a2_c4 = sheet_1["a2:c4"]
print(cell_a2_c4)

# Append a new row to the end of the sheet.
sheet_1.append([1004, 4, 11.89])

# Save the sheet.
wb.save("transaction2.xlsx")
